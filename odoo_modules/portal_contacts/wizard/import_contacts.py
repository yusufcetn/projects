# -*- coding: utf-8 -*-

from odoo import api, fields, models, _
from odoo.exceptions import UserError
from io import BytesIO
import openpyxl
import base64

class WizardImportContact(models.TransientModel):
    _name = 'wizard.import.contact'
    _description = 'Import Contacts'

    name = fields.Binary('File')
    msg = fields.Text('Messages', readonly=True)

    def import_contact(self, partner_type, user_id):
        fdata = self.name and base64.decodebytes(self.name) or False
        partner_obj = self.env['res.partner']
        country_obj = self.env['res.country']  # Ülke için model
        message_total = ""

        input = BytesIO(fdata)
        input.seek(4)  # Bazı binary formatlarında gerekebilir
        wb = openpyxl.load_workbook(input)
        sheet = wb._sheets[0]

        # İlk satırdaki başlıkları normalize et ve belirle
        headers = {}
        for col_num in range(1, sheet.max_column + 1):
            header_value = sheet.cell(row=1, column=col_num).value
            if header_value:
                # Başlıkları normalize ederek boşlukları ve büyük/küçük harfleri kaldırıyoruz
                normalized_header = header_value.strip().replace(" ", "_").lower()
                headers[normalized_header] = col_num

        for rowNum in range(2, sheet.max_row + 1):
            if not sheet.cell(row=rowNum, column=1).value:
                continue

            partner_data = {}

            # Ortak alanlar (educator_partner_id, partner_type, email, country)
            partner_data.update({
                'educator_partner_id': user_id.partner_id.id,
                'partner_type': partner_type,
                'user_id': user_id.partner_id.user_id.id if user_id.partner_id.user_id else False,
                'email': headers.get('email') and sheet.cell(row=rowNum, column=headers['email']).value or False,
            })

            # Country (ülke) için country tablosunda arama yapıyoruz
            if 'country' in headers:
                country_name = sheet.cell(row=rowNum, column=headers['country']).value
                if country_name:
                    country = country_obj.search([('name', '=', country_name)], limit=1)
                    if country:
                        partner_data['country_id'] = country.id
                    else:
                        message_total += f"Row {rowNum}: Country '{country_name}' not found.\n"
                        continue  # Eğer ülke bulunamazsa bu kaydı atla

            # Partner türüne göre ek alanlar
            if partner_type in ['student', 'teacher', 'subpartner']:
                # student, teacher ve subpartner için name
                partner_data['name'] = headers.get('name') and sheet.cell(row=rowNum,
                                                                          column=headers['name']).value or False

                # Öğrenci ve öğretmenler için ek alanlar
                if partner_type in ['student', 'teacher']:
                    partner_data.update({
                        'last_name': headers.get('last_name') and sheet.cell(row=rowNum, column=headers[
                            'last_name']).value or False,
                        'birth_date': headers.get('birth_date') and sheet.cell(row=rowNum, column=headers[
                            'birth_date']).value or False,
                        'user_gender': headers.get('gender') and sheet.cell(row=rowNum,
                                                                       column=headers['gender']).value.lower() or False,
                    })

                    # Öğrenci için öğretmen ID sorgulama
                    if partner_type == 'student':
                        teacher_id = headers.get('teacher_id') and sheet.cell(row=rowNum, column=headers['teacher_id']).value or False
                        if teacher_id:
                            teacher = partner_obj.search([
                                ('student_number', '=', teacher_id),
                                ('partner_type', '=', 'teacher'),
                            ], limit=1)
                            if teacher:
                                partner_data['teacher_id'] = teacher.id
                            else:
                                message_total += f"Row {rowNum}: Teacher with ID '{teacher_id}' not found'.\n"
                                continue


                # Teacher ve Subpartner için telefon
                if partner_type in ['teacher', 'subpartner']:
                    partner_data['phone'] = headers.get('phone') and sheet.cell(row=rowNum, column=headers['phone']).value or False

            # Okul işlemi
            if partner_type == 'school':
                school_name = sheet.cell(row=rowNum, column=headers['school_name']).value
                partner_data['name'] = school_name  # Sadece okul ID'si kullanılacak
                partner_data.update({
                    'is_company': True,
                    'phone': headers.get('phone') and sheet.cell(row=rowNum, column=headers['phone']).value or False,
                })

            # Öğrenci ve öğretmenler için var olan bir okulu bulma
            if partner_type in ['student', 'teacher']:
                school_id = sheet.cell(row=rowNum, column=headers['school_id']).value
                if school_id:
                    school = partner_obj.search([('student_number', '=', school_id), ('partner_type', '=', 'school')], limit=1)
                    if school:
                        partner_data['school_id'] = school.id
                    else:
                        message_total += f"Row {rowNum}: School with ID '{school_id}' not found for {partner_type}. Cannot proceed.\n"
                        continue

            # Partneri bul veya oluştur
            existing_partner = partner_obj.search([('email', '=', partner_data.get('email'))], limit=1)

            if existing_partner:
                existing_user = self.env['res.users'].search([('login', '=', partner_data.get('email'))], limit=1)

                if not existing_user:
                    user_vals = {
                        'partner_id': existing_partner.id,
                        'login': existing_partner.email,
                        'groups_id': [(6, 0, [self.env.ref('base.group_portal').id])]
                    }
                    user = self.env['res.users'].create(user_vals)

                    user.action_reset_password()
                else:
                    existing_partner.write(partner_data)
            else:
                existing_user = self.env['res.users'].search([('login', '=', partner_data.get('email'))], limit=1)

                if existing_user:
                    message_total += _("Row %d: A user with email '%s' already exists. Skipping import.\n") % (
                        rowNum, partner_data.get('email'))
                    continue

                new_partner = partner_obj.create(partner_data)

                user_vals = {
                    'partner_id': new_partner.id,
                    'login': new_partner.email,
                    'groups_id': [(6, 0, [self.env.ref('base.group_portal').id])]  # Portal grubunu ekliyoruz
                }
                user = self.env['res.users'].create(user_vals)

                user.action_reset_password()

        if message_total:
            raise UserError(_(message_total))